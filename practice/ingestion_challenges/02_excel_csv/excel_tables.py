"""
PATHOLOGY (a) + (f): One sheet is NOT one table.

Naive extraction (`pd.read_excel`) treats a worksheet as a single rectangular
frame. Real exports pack MULTIPLE logical tables on one sheet, separated by title
rows, blank spacer rows, and blank columns; they use MERGED cells for visual
grouping (a merged region stores its value only in the top-left cell, everything
else reads as None); and broken formulas flatten to error sentinels (#REF!,
#DIV/0!) that poison a numeric column.

This module, using only openpyxl (no pandas read_excel magic):
  - reads the sheet into a dense grid of cell values,
  - EXPANDS merged regions by forward-filling the anchor value into the block
    (so a merged group label repeats down its rows),
  - DETECTS table regions by scanning for all-blank separator rows,
  - trims title rows (a row with a single populated cell over an otherwise wide
    block) from the top of each region,
  - and QUARANTINES Excel error sentinels, distinguishing a cached error value
    from a live formula string via openpyxl's data_only vs formula load.

Interview line: "A worksheet is a canvas, not a table. Production ingestion first
segments the canvas into table regions, then normalizes each -- otherwise your
chunks straddle unrelated tables and retrieval returns nonsense."
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "fixtures", "dirty.xlsx")

ERROR_SENTINELS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


@dataclass
class TableRegion:
    """A rectangular block of rows on a sheet that forms one logical table."""
    top: int                 # 1-based first data/header row on the sheet
    bottom: int
    rows: list[list[Any]] = field(default_factory=list)
    title: str | None = None


def _read_grid(ws) -> list[list[Any]]:
    """Dense grid of *values* with merged regions forward-filled into the block."""
    max_r, max_c = ws.max_row, ws.max_column
    grid = [[ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
            for r in range(1, max_r + 1)]

    # Forward-fill every merged range with its anchor (top-left) value.
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_cc, max_rr = range_boundaries(str(mr))
        anchor = grid[min_r - 1][min_c - 1]
        for r in range(min_r, max_rr + 1):
            for c in range(min_c, max_cc + 1):
                grid[r - 1][c - 1] = anchor
    return grid


def _row_blank(row: list[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


def _populated_count(row: list[Any]) -> int:
    return sum(0 if (v is None or (isinstance(v, str) and v.strip() == "")) else 1
               for v in row)


def detect_regions(grid: list[list[Any]]) -> list[TableRegion]:
    """Split a grid into table regions on runs of all-blank rows.

    A title/banner row (a block of only single-populated rows) is not a table on
    its own -- it is carried forward as the title of the NEXT real table. This is
    why a standalone banner separated from its table by a spacer row still ends
    up attached, not emitted as an empty frame.
    """
    regions: list[TableRegion] = []
    cur: list[tuple[int, list[Any]]] = []
    pending_title: str | None = None

    def flush():
        nonlocal pending_title
        if not cur:
            return
        top = cur[0][0]
        bottom = cur[-1][0]
        body = [r for _, r in cur]

        # A region that is entirely single-populated rows is a banner block: hold
        # its text as a pending title for the next table and emit nothing now.
        if all(_populated_count(r) <= 1 for r in body):
            cell = next((v for r in body for v in r if v not in (None, "")), None)
            if cell is not None:
                pending_title = str(cell)
            return

        # Otherwise peel any leading single-populated title rows off the top.
        title = pending_title
        pending_title = None
        while body and _populated_count(body[0]) <= 1:
            cell = next((v for v in body[0] if v not in (None, "")), None)
            title = str(cell) if cell is not None else title
            body = body[1:]
            top += 1
        if len(body) >= 2:   # need at least a header + one data row
            regions.append(TableRegion(top=top, bottom=bottom, rows=body, title=title))

    for i, row in enumerate(grid, start=1):
        if _row_blank(row):
            flush()
            cur = []
        else:
            cur.append((i, row))
    flush()
    return regions


def region_to_df(region: TableRegion) -> pd.DataFrame:
    """First row of the region body is treated as the header.

    Column selection is POSITIONAL, not by label -- a two-row/merged header
    produces duplicate names ('Revenue','Revenue'), and selecting duplicate
    labels with df.loc would explode the frame width. We mask by index instead.
    """
    header = [("" if v is None else str(v)).strip() for v in region.rows[0]]
    data = [row[:len(header)] for row in region.rows[1:]]
    ncol = len(header)
    # keep a column if its header is non-empty OR any data cell is populated
    keep = []
    for j in range(ncol):
        named = header[j] != ""
        has_data = any(j < len(r) and r[j] not in (None, "") for r in data)
        keep.append(named or has_data)
    header = [header[j] for j in range(ncol) if keep[j]]
    data = [[r[j] for j in range(ncol) if keep[j]] for r in data]
    df = pd.DataFrame(data, columns=header)
    return df.reset_index(drop=True)


def scan_error_cells(path: str = XLSX) -> pd.DataFrame:
    """PATHOLOGY (f): quarantine error sentinels; classify formula vs cached value.

    openpyxl(data_only=False) gives the stored string ('=B2/B3' or '#REF!').
    openpyxl(data_only=True) gives the last cached computed value (None if the
    file was never opened by Excel). Comparing the two lets us label each cell.
    """
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    out = []
    for name in wb_f.sheetnames:
        wf, wv = wb_f[name], wb_v[name]
        for r in range(1, wf.max_row + 1):
            for c in range(1, wf.max_column + 1):
                raw = wf.cell(row=r, column=c).value
                cached = wv.cell(row=r, column=c).value
                is_formula = isinstance(raw, str) and raw.startswith("=")
                is_error = (isinstance(raw, str) and raw in ERROR_SENTINELS) or \
                           (isinstance(cached, str) and cached in ERROR_SENTINELS)
                if is_formula or is_error:
                    out.append({
                        "sheet": name, "cell": f"{r},{c}",
                        "raw": raw, "cached": cached,
                        "kind": "formula" if is_formula else "cached_error",
                        "quarantine": bool(is_error),
                    })
    return pd.DataFrame(out)


def _demo():
    print("=" * 70)
    print("EXCEL TABLE REGIONS + MERGED CELLS + ERROR SENTINELS")
    print("=" * 70)
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Q3_Report"]

    print("\n[BEFORE] raw pd.read_excel sees ONE ragged frame:")
    naive = pd.read_excel(XLSX, sheet_name="Q3_Report", header=None)
    print(naive.head(14).to_string())

    grid = _read_grid(ws)
    regions = detect_regions(grid)
    print(f"\n[AFTER] detected {len(regions)} logical table(s) on the sheet:")
    for i, reg in enumerate(regions, 1):
        print(f"\n  --- Region {i}  rows {reg.top}..{reg.bottom}  "
              f"title={reg.title!r}")
        df = region_to_df(reg)
        print(df.to_string(index=False))

    print("\n[MERGED CELLS] forward-filled group labels visible in region bodies "
          "above (no stray None from merged anchors).")

    print("\n[ERROR SENTINELS] quarantine report:")
    print(scan_error_cells().to_string(index=False))


if __name__ == "__main__":
    _demo()
