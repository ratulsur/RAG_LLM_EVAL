"""
Synthesize DIRTY spreadsheet fixtures so the handlers run with zero external files.

Produces:
  dirty.xlsx  -- multi-table sheet, title rows, blank spacers, merged group labels,
                 two-row header, units-in-headers, mixed-type columns, ambiguous dates,
                 Excel-serial dates, and formula/error cells (#REF!, #DIV/0!).
  dirty.csv   -- BOM + Latin-1 bytes, embedded commas/quotes, ';' delimiter,
                 sentinel nulls, trailing junk rows ("Total", sign-off).

Run:  python fixtures/make_fixtures.py     (writes both next to this file)
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "dirty.xlsx")
CSV = os.path.join(HERE, "dirty.csv")


def make_xlsx(path: str = XLSX) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3_Report"

    # ---- Table 1: title row + merged group label + two-row header + dirty body ----
    ws["A1"] = "APAC Regional Sales -- Q3 FY24 (auto-exported)"   # title row (row 1)
    ws["A1"].font = Font(bold=True, size=14)
    # row 2 blank spacer (left empty)

    # Two-row header spanning A..E starting row 3.
    # Row 3 = group labels (merged cells), Row 4 = leaf labels.
    ws["A3"] = "Region"                 # merged A3:A4 (row label, vertical merge)
    ws.merge_cells("A3:A4")
    ws["B3"] = "Revenue"                # merged B3:C4 group over two leaf cols? keep simple: B3:C3
    ws.merge_cells("B3:C3")
    ws["D3"] = "Ops"
    ws.merge_cells("D3:E3")
    # leaf header (row 4)
    ws["B4"] = "Revenue (INR Cr)"
    ws["C4"] = "growth_pct"
    ws["D4"] = "Headcount"
    ws["E4"] = "Attrition (%)"

    # Body rows 5..8 -- deliberately mixed types / dirty
    body1 = [
        ["North", "1,240", "12.5", 340, "8.2"],
        ["South", "N/A", "9", "TBD", "6.1"],           # sentinel + text-number
        ["East", "980", "-", "pending", "-"],           # '-' sentinel + un-coercible
        ["West", "INR 1,110", "15.0", 275, "7.7"],      # currency-prefixed number
    ]
    r = 5
    for row in body1:
        for c, val in enumerate(row):
            ws.cell(row=r, column=c + 1, value=val)
        r += 1

    # blank spacer rows 9,10 (left empty) -> table separator

    # ---- Table 2 (SECOND logical table on SAME sheet), starts row 11 ----
    ws["A11"] = "Contract Milestones"                    # title row for table 2
    ws["A11"].font = Font(bold=True)
    # header row 12
    hdr2 = ["Contract ID", "Signed Date", "Amount (USD)", "Status"]
    for c, val in enumerate(hdr2):
        ws.cell(row=12, column=c + 1, value=val)
    # body 13..16 -- AMBIGUOUS dates (DD/MM vs MM/DD mixed), 2-digit year, serials
    body2 = [
        ["C-001", "03/04/2024", "$12,500.00", "Active"],   # ambiguous 03/04
        ["C-002", "13/05/2024", "$8,000", "Active"],        # 13 -> DD/MM certain
        ["C-003", "07/22/23", "1500", "Closed"],           # MM/DD/2-digit-year
        ["C-004", 45566, "$3,200", "Active"],              # Excel serial date
    ]
    r = 13
    for row in body2:
        for c, val in enumerate(row):
            ws.cell(row=r, column=c + 1, value=val)
        r += 1

    # ---- Table 3: formula / error cells on a second sheet ----
    ws2 = wb.create_sheet("Calcs")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    # write literal error strings -- mimics an export that flattened broken formulas
    calc = [
        ["margin", "=B2/B3"],       # a real formula (openpyxl stores the string)
        ["ratio", "#DIV/0!"],       # cached error value
        ["ref", "#REF!"],           # broken reference
        ["ok", 42],
    ]
    r = 2
    for name, val in calc:
        ws2.cell(row=r, column=1, value=name)
        ws2.cell(row=r, column=2, value=val)
        r += 1

    wb.save(path)
    return path


def make_csv(path: str = CSV) -> str:
    """Write raw bytes: UTF-8 BOM header line, then a Latin-1 encoded body with
    embedded commas/quotes, a ';'-heavy structure, sentinels, and junk trailer."""
    bom = b"\xef\xbb\xbf"  # UTF-8 BOM
    # Header + rows. Note: this file uses ',' as delimiter but fields contain commas
    # inside quotes, plus one Latin-1 non-ASCII byte (0xe9 = 'e-acute' in Latin-1).
    lines = []
    lines.append(b'Client,Note,Amount,Signed')
    # embedded comma inside quotes + embedded quote (doubled per RFC4180)
    lines.append(b'"Acme, Inc.","Renewal ""priority"" account","1,240",03/04/2024')
    # Latin-1 accented byte in a name (Andr\xe9)  -> will break naive utf-8 decode
    lines.append(b'Andr\xe9 & Co,Standard,980,13/05/2024')
    lines.append(b'"Beta LLC","note; with semicolon",N/A,07/22/23')
    # trailing junk rows a human appended
    lines.append(b'Total,,2220,')
    lines.append(b'Signed off by: R. Sur -- 2024-10-01,,,')
    blob = bom + b"\r\n".join(lines) + b"\r\n"
    with open(path, "wb") as f:
        f.write(blob)
    return path


if __name__ == "__main__":
    print("wrote:", make_xlsx())
    print("wrote:", make_csv())
